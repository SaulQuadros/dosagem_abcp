
from __future__ import annotations
import pandas as pd
from typing import Dict, Any, Optional, Tuple, List
from openpyxl import load_workbook

def _win(ws, r, c, h, w):
    return [[ws.cell(row=rr, column=cc).value for cc in range(c, c+w)] for rr in range(r, r+h)]

def load_abcp_tables(excel_path) -> Dict[str, Any]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['ABCP'] if 'ABCP' in wb.sheetnames else wb.active

    # Known anchors from inspection
    t1 = pd.DataFrame(_win(ws, 1, 1, 12, 12))      # includes Tabela 1 area
    t2 = pd.DataFrame(_win(ws, 11, 1, 16, 16))     # includes Tabela 2 + 3
    t4 = pd.DataFrame(_win(ws, 2, 8, 16, 12))      # Tabela 4
    t5 = pd.DataFrame(_win(ws, 19, 8, 16, 10))     # Tabela 5 (britas)

    # --- Parse Tabela 1 ---
    # Headers: row 3 contains I, II, III, IV (0-based index 3 -> row 4 human)
    classes = [str(x) for x in t1.iloc[3,2:6].tolist()]
    # a/c max: rows 4 (CA) and 5 (CP)
    ac_CA = t1.iloc[4,2:6].astype(float).tolist()
    ac_CP = t1.iloc[5,2:6].astype(float).tolist()
    # fck minimo (NBR8953): rows 6 (CA) and 7 (CP) e.g., C20,C25 strings -> parse numeric
    fck_CA = [int(str(s).replace('C','')) for s in t1.iloc[6,2:6].tolist()]
    fck_CP = [int(str(s).replace('C','')) for s in t1.iloc[7,2:6].tolist()]
    # consumo mínimo de cimento: row 9 (0-based index 8? from earlier, but reading rows 7..)
    cc_row = t1.iloc[9-1, 2:6] if not pd.isna(t1.iloc[9-1, 2]) else t1.iloc[8,2:6]
    cc_min = [float(x) for x in cc_row.tolist()]

    tabela1 = {
        "classes": classes,  # e.g. ["I","II","III","IV"]
        "ac_max": {"CA": ac_CA, "CP": ac_CP},
        "fck_min": {"CA": fck_CA, "CP": fck_CP},
        "cc_min": cc_min
    }

    # --- Parse Tabela 2 (Ca) ---
    # Dmax row is at t2.iloc[2,1:] e.g. [9.5, 19, 25, 32, 38]
    dmax_labels = [x for x in t2.iloc[2,1:].tolist() if x is not None]
    # Slump rows start at row 4..6; first col has "40-60", etc.
    slump_labels = [str(t2.iloc[i,0]) for i in [4,5,6]]
    ca_rows = []
    for i in [4,5,6]:
        row = [t2.iloc[i, j] for j in range(1, 1+len(dmax_labels))]
        ca_rows.append(row)
    tabela2 = {"dmax": dmax_labels, "slump": slump_labels, "ca": ca_rows}

    # --- Parse Tabela 3 (Vb %) ---
    # Dmax headers at row 10, cols 1..5; MF at col 0 from row 12..15
    dmax3 = [t2.iloc[10, j] for j in range(1,6)]
    mf_list = [t2.iloc[i,0] for i in range(12,16)]
    vb = [[t2.iloc[i,j] for j in range(1,6)] for i in range(12,16)]
    tabela3 = {"dmax": dmax3, "mf": mf_list, "vb": vb}

    # --- Tabela 4 (Sd by Condição A/B/C) ---
    sd_vals = {"A": float(t4.iloc[3,0]), "B": float(t4.iloc[3,1]), "C": float(t4.iloc[3,2])}
    tabela4 = {"sd": sd_vals}

    # --- Tabela 5 (Britas info) ---
    britas = {}
    for i in range(3,8):
        name = t5.iloc[i,0]
        if name:
            britas[str(name)] = {"dmax": t5.iloc[i,1], "faixa": t5.iloc[i,2]}
    tabela5 = britas

    return {"tabela1": tabela1, "tabela2": tabela2, "tabela3": tabela3, "tabela4": tabela4, "tabela5": tabela5}

def lookup_ca_from_tables(tables, dmax, slump) -> Optional[float]:
    t2 = tables["tabela2"]
    d_list = t2["dmax"]; s_list = t2["slump"]; ca = t2["ca"]
    try:
        i = s_list.index(str(slump))
        j = d_list.index(dmax)
        return float(ca[i][j])
    except Exception:
        return None

def lookup_vb_from_tables(tables, mf, dmax) -> Optional[float]:
    t3 = tables["tabela3"]
    d_list = t3["dmax"]; mf_list = t3["mf"]; vb = t3["vb"]
    # exact match preferred
    try:
        i = mf_list.index(mf)
        j = d_list.index(dmax)
        return float(vb[i][j])
    except Exception:
        # nearest MF fallback
        try:
            diffs = [abs((m or 0)-mf) for m in mf_list]
            i = int(diffs.index(min(diffs)))
            j = d_list.index(dmax)
            return float(vb[i][j])
        except Exception:
            return None

def lookup_limits_from_tabela1(tables, tipo, classe) -> Dict[str, float]:
    t1 = tables["tabela1"]
    classes = t1["classes"]
    j = classes.index(str(classe))
    ac_max = float(t1["ac_max"][tipo][j])
    fck_min = float(t1["fck_min"][tipo][j])
    cc_min = float(t1["cc_min"][j])
    return {"ac_max": ac_max, "fck_min": fck_min, "cc_min": cc_min}

def lookup_sd_from_tabela4(tables, cond) -> float:
    return float(tables["tabela4"]["sd"][cond])
