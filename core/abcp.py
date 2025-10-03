
from __future__ import annotations
import pandas as pd
import numpy as np
from typing import Dict, Any, Union, Optional, Tuple, List
from pathlib import Path
from openpyxl import load_workbook

def _read_named_range(wb, name: str):
    # Return (sheet_name, list of (row, col)) for a defined name
    try:
        dn = wb.defined_names.get(name)
        if dn is None:
            return None
        dests = list(dn.destinations)
        if not dests:
            return None
        sheet_name, ref = dests[0]
        from openpyxl.utils import range_boundaries, coordinate_to_tuple
        cells = []
        if ":" in ref:
            min_col, min_row, max_col, max_row = range_boundaries(ref)
            for r in range(min_row, max_row+1):
                for c in range(min_col, max_col+1):
                    cells.append((r, c))
        else:
            r, c = coordinate_to_tuple(ref)
            cells.append((r, c))
        return (sheet_name, cells)
    except Exception:
        return None

def _extract_matrix(wb, name: str):
    nr = _read_named_range(wb, name)
    if nr is None:
        return None
    sheet_name, cells = nr
    ws = wb[sheet_name]
    rows = sorted({r for r,c in cells})
    cols = sorted({c for r,c in cells})
    data = []
    for r in rows:
        row_vals = []
        for c in cols:
            row_vals.append(ws.cell(row=r, column=c).value)
        data.append(row_vals)
    df = pd.DataFrame(data)
    return df

def load_ca_lookup(excel_path: Union[str, bytes, Path]):
    # Load Ca lookup from named ranges: Tabela2Ca, Tabela2Dmax, Tabela2Slump
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception:
        return None
    df_ca = _extract_matrix(wb, "Tabela2Ca")
    df_dmax = _extract_matrix(wb, "Tabela2Dmax")
    df_slump = _extract_matrix(wb, "Tabela2Slump")
    if df_ca is None or df_dmax is None or df_slump is None:
        return None
    dmax_vals = [v for v in df_dmax.values.flatten().tolist() if v is not None]
    slump_vals = [v for v in df_slump.values.flatten().tolist() if v is not None]
    dmax_labels = [str(v).strip() for v in dmax_vals]
    slump_labels = [str(v).strip() for v in slump_vals]
    ca = df_ca.replace("", np.nan).astype(float)
    if ca.shape != (len(dmax_labels), len(slump_labels)):
        if ca.T.shape == (len(dmax_labels), len(slump_labels)):
            ca = ca.T
        else:
            return None
    return {"dmax_labels": dmax_labels, "slump_labels": slump_labels, "ca_table": ca.values}

def lookup_ca(ca_lookup: Dict[str, Any], dmax_label: str, slump_label: str):
    try:
        d_list = ca_lookup["dmax_labels"]
        s_list = ca_lookup["slump_labels"]
        table = ca_lookup["ca_table"]
        i = d_list.index(str(dmax_label).strip())
        j = s_list.index(str(slump_label).strip())
        return float(table[i, j])
    except Exception:
        return None

def compute_abcp(
    ac: float,
    Ca_L: float,
    rho_w: float,
    Cc_min: float,
    rho_c: float,
    rho_s_grain: float,
    rho_b_menor: float,
    rho_b_maior: float,
    Cb_total: float,
    perc_b_menor: int,
    U_areia: float,
    I_inch: float,
    rho_s_bulk: float,
    a_areia: float = 0.0,
    a_brita: float = 0.0,
    U_brita: float = 0.0,
) -> Dict[str, Any]:
    # Água de dosagem (massa) a partir de Ca em L/m³ e rho_w
    P33 = Ca_L * (rho_w / 1000.0)

    # Consumo de cimento
    Cc_calc = P33 / max(ac, 1e-9)
    Cc = max(Cc_calc, Cc_min)

    # Distribuição de brita
    frac_menor = np.clip(perc_b_menor/100.0, 0, 1)
    Cb_menor = Cb_total * frac_menor
    Cb_maior = Cb_total * (1.0 - frac_menor)

    # Volumes absolutos
    V_c = Cc / rho_c
    V_w = P33 / rho_w
    V_g_menor = Cb_menor / rho_b_menor
    V_g_maior = Cb_maior / rho_b_maior
    V_g_total = V_g_menor + V_g_maior

    Vm = 1.0 - (V_c + V_w + V_g_total)
    Vm = max(Vm, 0.0)

    # Areia (massa seca e úmida)
    Cm_seca = Vm * rho_s_grain
    U = U_areia / 100.0
    Cm_umida = Cm_seca * (1.0 + U)
    agua_areia_total = Cm_umida - Cm_seca

    # Brita
    Ub = U_brita / 100.0
    Cb_seco_total = Cb_total
    agua_brita_total = Cb_seco_total * Ub

    # Absorções
    a_s = a_areia / 100.0
    a_b = a_brita / 100.0
    agua_absorcao = (a_s * Cm_seca) + (a_b * Cb_seco_total)

    # Água de umidade
    agua_moist_total = agua_areia_total + agua_brita_total

    # Água efetiva a adicionar
    agua_adicionar_kg = P33 + agua_absorcao - agua_moist_total

    # Volumetria para obra (areia aparente + inchamento)
    V_areia_seca_aparente = Cm_seca / rho_s_bulk
    I = I_inch / 100.0
    V_areia_med_m3 = V_areia_seca_aparente * (1.0 + I)
    V_areia_med_L = V_areia_med_m3 * 1000.0

    return dict(
        P33=P33, Cc=Cc,
        Cb_menor=Cb_menor, Cb_maior=Cb_maior,
        V_c=V_c, V_w=V_w, V_g_total=V_g_total, Vm=Vm,
        Cm_seca=Cm_seca, Cm_umida=Cm_umida,
        agua_areia_total=agua_areia_total,
        agua_brita_total=agua_brita_total,
        agua_absorcao=agua_absorcao,
        agua_moist_total=agua_moist_total,
        agua_adicionar_kg=agua_adicionar_kg,
        V_areia_med_m3=V_areia_med_m3, V_areia_med_L=V_areia_med_L
    )

def load_tables_preview(excel_path: Union[str, bytes, Path]):
    previews = {}
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
        for name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=name, header=None, nrows=60, usecols="A:Z", engine="openpyxl")
            previews[name] = df
            if len(previews) >= 3:
                break
    except Exception as e:
        previews["Aviso"] = pd.DataFrame({"Mensagem": [f"Não foi possível ler o Excel: {e}"]})
    return previews
