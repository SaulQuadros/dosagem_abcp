
import streamlit as st
import pandas as pd
from pathlib import Path
from core.abcp import compute_abcp
from core.abcp_tables import (
    load_abcp_tables,
    lookup_ca_from_tables,
    lookup_vb_from_tables,
    lookup_limits_from_tabela1,
    lookup_sd_from_tabela4
)
from core.pdf_utils import generate_traco_pdf

st.set_page_config(page_title="Dosagem ABCP - Concreto", page_icon="🧮", layout="wide")

# --- Sidebar: identificação + Excel ---
with st.sidebar:
    st.header("📋 Identificação")
    projeto = st.text_input("Nome do Projeto", value="Obra A - Laje")
    tecnico = st.text_input("Técnico Responsável", value="Eng. Fulano de Tal")
    uso = st.selectbox("Tipo de Uso do Concreto", ["Estrutural", "Pavimento", "Viga/Laje", "Fundação", "Outros"], index=0)
    fabricado_em = st.selectbox("Fabricado em", ["Usina", "Canteiro"], index=0)
    st.markdown("---")
    st.subheader("📄 Excel com Tabelas (recomendado)")
    excel_file = st.file_uploader("Carregar Excel 04_Dosagem_Concreto_ABCP.xlsx", type=["xlsx"], key="uploader_sidebar")
    st.caption("O app usa a aba 'ABCP' para ler as Tabelas 1..5 (consulta e cálculos).")

st.title("🧮 Dosagem de Concreto — Método ABCP")
tab_calc, tab_tabelas = st.tabs(["🧪 Dosagem ABCP", "📚 Tabela Consulta (ABCP)"])

# Carregar tabelas do Excel
tables = None
base_excel_path = Path("data/04_Dosagem_Concreto_ABCP.xlsx")
excel_path = None
if excel_file is not None:
    excel_path = excel_file
elif base_excel_path.exists():
    excel_path = str(base_excel_path)

if excel_path is not None:
    try:
        tables = load_abcp_tables(excel_path)
    except Exception as e:
        st.warning(f"Não foi possível ler as Tabelas do Excel: {e}")

# --- Aba 1: Dosagem (entradas manuais + resultados por tabelas) ---
with tab_calc:
    if not tables:
        st.info("Carregue o Excel para habilitar as escolhas dirigidas pelas Tabelas 1–5.")
        st.stop()

    # Build dynamic choices from tables
    classes_opts = tables["tabela1"]["classes"]
    tipos_opts = ["CA","CP"]
    cond_opts = list(tables["tabela4"]["sd"].keys())
    dmax_opts = tables["tabela2"]["dmax"]
    slump_opts = tables["tabela2"]["slump"]
    mf_opts = tables["tabela3"]["mf"]

    st.subheader("1) Escolhas dirigidas pelas tabelas")
    colA, colB, colC = st.columns(3)
    with colA:
        tipo = st.radio("Tipo de Concreto (Tabela 1)", tipos_opts, horizontal=True, index=0)
        classe = st.selectbox("Classe de Agressividade (Tabela 1)", classes_opts, index=0)
        cond = st.radio("Condição de Preparo (Tabela 4)", cond_opts, horizontal=True, index=0)
    with colB:
        dmax = st.selectbox("Dmáx Agregado Graúdo (Tabela 2/5)", dmax_opts, index=0)
        slump = st.selectbox("Abatimento - Slump (Tabela 2)", slump_opts, index=0)
        # MF é manual, mas mostramos opções de referência (Tabela 3)
    with colC:
        perc_b_menor = st.slider("% Brita Menor", 0, 100, 50, step=5)
        ac = st.number_input("Fator a/c (projeto)", min_value=0.30, max_value=0.75, value=0.45, step=0.01, format="%.2f")

    # Lookups derivados das tabelas
    limits = lookup_limits_from_tabela1(tables, tipo, classe)     # a/c max, fck_min, cc_min
    sd = lookup_sd_from_tabela4(tables, cond)                     # Sd
    Ca_L_tab = lookup_ca_from_tables(tables, dmax, slump)         # Ca (L/m³)

    st.markdown("---")
    st.subheader("2) Entradas NÃO calculadas por tabelas (preencher)")
    colD, colE, colF = st.columns(3)
    with colD:
        MF = st.number_input("Módulo de Finura da Areia (MF)", min_value=1.50, max_value=3.20, value=float(mf_opts[0]) if mf_opts else 2.20, step=0.05, format="%.2f")
        U_areia = st.number_input("Umidade da Areia — % (base seca)", 0.0, 12.0, 6.0, 0.1, format="%.1f")
        I_inch = st.number_input("Inchamento da Areia — %", 0.0, 35.0, 20.0, 1.0, format="%.0f")
    with colE:
        rho_w = st.number_input("Massa específica da Água — kg/m³", 980.0, 1020.0, 1000.0, 1.0, format="%.0f")
        rho_c = st.number_input("Massa específica do Cimento (grão) — kg/m³", 2900.0, 3300.0, 3100.0, 10.0, format="%.0f")
        rho_s_grain = st.number_input("Massa específica da Areia (grão) — kg/m³", 2500.0, 2700.0, 2650.0, 10.0, format="%.0f")
    with colF:
        rho_s_bulk = st.number_input("Massa unitária aparente da Areia (seca) — kg/m³", 1200.0, 1800.0, 1470.0, 10.0, format="%.0f")
        a_areia = st.number_input("Absorção da Areia — %", 0.0, 5.0, 0.0, 0.1, format="%.1f")
        U_brita = st.number_input("Umidade da Brita — % (base seca)", 0.0, 6.0, 0.0, 0.1, format="%.1f")

    colG, colH, colI = st.columns(3)
    with colG:
        rho_b_grain_menor = st.number_input("Massa específica Brita Menor (grão) — kg/m³", 2600.0, 2800.0, 2700.0, 10.0, format="%.0f")
        rho_b_grain_maior = st.number_input("Massa específica Brita Maior (grão) — kg/m³", 2600.0, 2800.0, 2700.0, 10.0, format="%.0f")
    with colH:
        rho_b_bulk_menor = st.number_input("Massa unitária aparente Brita Menor — kg/m³", 1200.0, 1800.0, 1430.0, 10.0, format="%.0f")
        rho_b_bulk_maior = st.number_input("Massa unitária aparente Brita Maior — kg/m³", 1200.0, 1800.0, 1430.0, 10.0, format="%.0f")
    with colI:
        rho_b_bulk_media = st.number_input("Massa unitária aparente Brita (média) — kg/m³", 1200.0, 1800.0, 1500.0, 10.0, format="%.0f")
        a_brita = st.number_input("Absorção da Brita — % (total)", 0.0, 5.0, 1.0, 0.1, format="%.1f")

    # Resultados provenientes das Tabelas
    st.markdown('---')
    st.subheader("3) Resultados provenientes das Tabelas (automáticos)")

    if Ca_L_tab is None:
        st.error("Não foi possível obter Ca (L/m³) a partir da Tabela 2. Verifique o Excel/seleções.")
        Ca_L = 200.0
    else:
        Ca_L = float(Ca_L_tab)

    Vb = lookup_vb_from_tables(tables, MF, dmax)
    if Vb is None:
        st.error("Não foi possível obter Vb (Tabela 3). Verifique MF/Dmáx e o Excel.")
        Vb = 0.70

    colR0, colR1, colR2, colR3 = st.columns(4)
    with colR0:
        st.metric("Ca (L/m³) — Tabela 2", f"{Ca_L:.1f}")
    with colR1:
        st.metric("a/c máximo — Tabela 1", f"{limits['ac_max']:.2f}")
        if ac > limits["ac_max"]:
            st.error("a/c acima do máximo permitido (Tabela 1).")
    with colR2:
        st.metric("Cc mínimo — Tabela 1 (kg/m³)", f"{limits['cc_min']:.0f}")
        st.metric("fck mínimo — Tabela 1", f"C{int(limits['fck_min'])}")
    with colR3:
        st.metric("Sd — Tabela 4 (MPa)", f"{sd:.2f}")
        st.metric("fck alvo (28d)", f"{(limits['fck_min'] + 1.65*sd):.1f}")

    # --- Cálculo do traço (1 m³) ---
    out = compute_abcp(
        ac=ac, Ca_L=Ca_L, rho_w=rho_w, Cc_min=limits["cc_min"],
        rho_c=rho_c, rho_s_grain=rho_s_grain,
        rho_b_menor=rho_b_grain_menor, rho_b_maior=rho_b_grain_maior,
        Cb_total=0.0, perc_b_menor=perc_b_menor,
        U_areia=U_areia, I_inch=I_inch, rho_s_bulk=rho_s_bulk,
        a_areia=a_areia, a_brita=a_brita, U_brita=U_brita
    )

    # Cálculo de Cb pela lógica da planilha (Vb × massas unitárias aparentes)
    frac_menor = perc_b_menor/100.0
    if perc_b_menor in (0, 100):
        Cb_total = Vb * rho_b_bulk_media
        Cb_menor = Cb_total if perc_b_menor == 100 else 0.0
        Cb_maior = 0.0 if perc_b_menor == 100 else Cb_total
    else:
        Cb_menor = Vb * rho_b_bulk_menor * frac_menor
        Cb_maior = Vb * rho_b_bulk_maior * (1.0 - frac_menor)
        Cb_total = Cb_menor + Cb_maior

    # Recalcular volumes das britas com densidade de grão
    V_g_total = Cb_menor / rho_b_grain_menor + Cb_maior / rho_b_grain_maior

    # Recalcular volumes e Vm
    V_w = out["P33"] / rho_w
    V_c = out["Cc"] / rho_c
    Vm = max(1.0 - (V_c + V_w + V_g_total), 0.0)

    # Areia
    Cm_seca = Vm * rho_s_grain
    U = U_areia / 100.0
    Cm_umida = Cm_seca * (1.0 + U)
    agua_areia = Cm_umida - Cm_seca

    # Águas brita/absorção
    agua_brita = (U_brita/100.0) * (Cb_menor + Cb_maior)
    agua_abs = (a_areia/100.0)*Cm_seca + (a_brita/100.0)*(Cb_menor+Cb_maior)
    agua_adicionar_kg = out["P33"] + agua_abs - (agua_areia + agua_brita)

    # Volumetria de obra (areia aparente + inchamento)
    V_areia_seca_aparente = Cm_seca / rho_s_bulk
    V_areia_med_m3 = V_areia_seca_aparente * (1 + I_inch/100.0)
    V_areia_med_L = V_areia_med_m3 * 1000.0

    st.markdown('---')
    st.subheader("4) Resultados (por 1 m³)")
    colX1, colX2 = st.columns([1.2,1])
    with colX1:
        st.markdown("**Massas (kg/m³)**")
        st.table(pd.DataFrame({
            "Grandeza": [
                "Água (massa alvo) — P33",
                "Cimento — Cc",
                "Brita Menor — Cb,menor",
                "Brita Maior — Cb,maior",
                "Total Britas — Cb",
                "Areia (seca) — Cm,seca",
                "Areia (úmida) — Cm,úmida",
                "Água na areia (umidade)",
                "Água na brita (umidade)",
                "Água absorvida (areia+brita)",
                "Água a adicionar (kg/m³)"
            ],
            "Valor (kg/m³)": [
                round(out["P33"],2),
                round(out["Cc"],2),
                round(Cb_menor,2),
                round(Cb_maior,2),
                round(Cb_total,2),
                round(Cm_seca,2),
                round(Cm_umida,2),
                round(agua_areia,2),
                round(agua_brita,2),
                round(agua_abs,2),
                round(agua_adicionar_kg,2),
            ]
        }))
    with colX2:
        st.markdown("**Volumes (m³ e L)**")
        st.table(pd.DataFrame({
            "Grandeza": [
                "Volume do cimento — Vc (m³)",
                "Volume da água — Vw (m³)",
                "Volume das britas — Vg (m³)",
                "Volume de areia — Vm (m³)",
                "Areia a medir (com inchamento) — m³",
                "Areia a medir — L"
            ],
            "Valor": [
                round(V_c,4),
                round(V_w,4),
                round(V_g_total,4),
                round(Vm,4),
                round(V_areia_med_m3,4),
                round(V_areia_med_L,1),
            ]
        }))

    st.markdown('---')
    st.subheader("5) Gerar PDF do Traço")
    ident = dict(projeto=projeto, tecnico=tecnico, uso=uso, fabricado_em=fabricado_em)
    inputs = dict(ac=ac, Ca_L=Ca_L, rho_w=rho_w, Cc_min=limits["cc_min"], rho_c=rho_c, rho_s_grain=rho_s_grain,
                  rho_b_menor=rho_b_grain_menor, rho_b_maior=rho_b_grain_maior, Cb_total=Cb_total, perc_b_menor=perc_b_menor,
                  U_areia=U_areia, I_inch=I_inch, rho_s_bulk=rho_s_bulk, a_areia=a_areia, a_brita=a_brita, U_brita=U_brita, MF=MF,
                  dmax=dmax, slump=slump, tipo=tipo, classe=classe, cond=cond)
    outputs = dict(P33=out["P33"], Cc=out["Cc"], Cb_menor=Cb_menor, Cb_maior=Cb_maior, V_c=V_c, V_w=V_w,
                   V_g_total=V_g_total, Vm=Vm, Cm_seca=Cm_seca, Cm_umida=Cm_umida, agua_areia_total=agua_areia,
                   agua_brita_total=agua_brita, agua_absorcao=agua_abs, agua_moist_total=agua_areia+agua_brita,
                   agua_adicionar_kg=agua_adicionar_kg, V_areia_med_m3=V_areia_med_m3, V_areia_med_L=V_areia_med_L)

    if st.button("Gerar PDF"):
        pdf_path = Path("traco_abcp.pdf")
        generate_traco_pdf(str(pdf_path), ident, inputs, outputs)
        with open(pdf_path, "rb") as f:
            st.download_button("Baixar PDF", f, file_name="traco_abcp.pdf", mime="application/pdf")

# --- Aba 2: Tabela Consulta ---
with tab_tabelas:
    st.subheader("Tabelas 1 a 5 (somente leitura) — aba 'ABCP' do Excel")
    if tables is None:
        st.info("Carregue o Excel na barra lateral para ver as tabelas.")
    else:
        import pandas as pd
        from openpyxl import load_workbook
        def to_df(ws, r,c,h,w):
            return pd.DataFrame([[ws.cell(row=rr, column=cc).value for cc in range(c, c+w)] for rr in range(r, r+h)])
        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb['ABCP'] if 'ABCP' in wb.sheetnames else wb.active
            st.markdown("**Tabela 1**")
            st.dataframe(to_df(ws, 1, 1, 12, 12), use_container_width=True, height=320)
            st.markdown("**Tabela 2**")
            st.dataframe(to_df(ws, 11, 1, 12, 12), use_container_width=True, height=320)
            st.markdown("**Tabela 3**")
            st.dataframe(to_df(ws, 19, 1, 12, 12), use_container_width=True, height=320)
            st.markdown("**Tabela 4**")
            st.dataframe(to_df(ws, 2, 8, 12, 12), use_container_width=True, height=240)
            st.markdown("**Tabela 5**")
            st.dataframe(to_df(ws, 19, 8, 12, 10), use_container_width=True, height=240)
        except Exception as e:
            st.warning(f"Falha ao exibir tabelas: {e}")
